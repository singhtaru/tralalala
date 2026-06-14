"""Convert the Amazon Now Excel catalog into frontend JSON."""

import json
import re
from pathlib import Path

from openpyxl import load_workbook


FRONTEND_ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = FRONTEND_ROOT / "data" / "source" / "AmazonNow_Dummy_Product_Dataset.xlsx"
OUTPUT_FILE = FRONTEND_ROOT / "src" / "data" / "products.json"

CATEGORY_IDS = {
    "Vegetables & Fruits": "fruits",
    "Atta Rice & Dal": "atta",
    "Oil Ghee & Masala": "oil",
    "Dairy Bread & Eggs": "dairy",
    "Bakery & Biscuits": "bakery",
    "Dry Fruits & Cereals": "cereal",
    "Chicken Meat & Fish": "meat",
    "Kitchenware": "kitchen",
    "Chips & Namkeen": "chips",
    "Sweets & Chocolates": "chocolate",
    "Drinks & Juices": "drinks",
    "Tea Coffee & More": "tea",
    "Instant Food": "instant",
    "Protein & Fitness": "health",
    "Ice Creams & Desserts": "icecream",
}


def extract_quantity(product_name: str) -> str:
    match = re.search(
        r"(\d+(?:\.\d+)?\s?(?:kg|g|ml|l)|pack of \d+)",
        product_name,
        re.IGNORECASE,
    )
    return match.group(1) if match else "1 pack"


def main() -> None:
    workbook = load_workbook(SOURCE_FILE, data_only=True, read_only=True)
    worksheet = workbook["AmazonNowProducts"]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    products = []

    for values in rows[1:]:
        row = dict(zip(headers, values))
        product_name = str(row["Product_Name"])
        products.append(
            {
                "id": row["Product_ID"],
                "name": product_name,
                "category": CATEGORY_IDS[row["Category"]],
                "categoryName": row["Category"],
                "price": row["Price_INR"],
                "rating": row["Rating"],
                "deliveryMins": row["Delivery_Time_Min"],
                "quantity": extract_quantity(product_name),
            }
        )

    OUTPUT_FILE.write_text(
        json.dumps(products, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    print(f"Generated {len(products)} products at {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
